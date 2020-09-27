from selenium import webdriver
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import cell
from openpyxl.styles import Alignment #, named_styles
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from selenium.webdriver.common.by import By
import sys


print('''
Программа предназначена для поиска ссылок на филиалы организации (точку продаж).
Вставьте прямую ссылку на страницу с филиалами организации без пробелов и других знаков. Пример: 'https://2gis.kz/almaty/branches/9429948590733484'
Если Ctrl+V в терминале не работает, попробуйте нажать правую кнопку мыши в месте для вставки после того, как скопировали ссылку.

У Вас должен быть устанолен браузер Firefox, потому что программа взаимодействует с ним для поиска дынныъ. 
Чтобы управлять Firefox с помощью программного кода используется 'geckodriver' https://github.com/mozilla/geckodriver/releases.
Если браузер не работает возможно папку с geckodriver.exe нужно добавить в PATH.
Версия geckodriver.exe должна соответсвовать версии Firefox на компьютере.

Файл будет сохранен в ту же папку, где была запущена программа.
''')
direct_link_to_branches = input('Введите ссылку, нажмите Enter: ')
city_name = direct_link_to_branches.split('/')
# print(city_name)
browser = webdriver.Firefox()

# Provide branch url 
# KDL OLYMP BRANCH url in Almaty example: 'https://2gis.kz/almaty/branches/9429948590733484'
# Aktau https://2gis.kz/aktau/branches/70000001028456934
# Astana https://2gis.kz/nur_sultan/branches/70000001018099222
browser.get(direct_link_to_branches)
branch_count = browser.find_elements(By.CLASS_NAME, "_1p8iqzw")[1].text

# LOOPS UNTIL EACH ELEMENT CONTAINING COMPANY LINK IS FOUND
while True:
    try:
        
        content_blocks = browser.find_elements_by_class_name('_vhuumw') 
        browser.execute_script("arguments[0].scrollIntoView();"  , content_blocks[-1] )
        #code returns error if  button not found on page
        add_company_button = browser.find_element_by_class_name('_4hzbziy')
         
    except:
         continue
    else:
        # scroll to add_company_button  
        browser.execute_script("arguments[0].scrollIntoView();"  , add_company_button )
        content_blocks = browser.find_elements_by_class_name('_vhuumw') 
        break

# # Scrolls part of page with addresses certain number of times
# for item in range(1,10):
#     content_blocks = browser.find_elements_by_class_name('_vhuumw') # _oqoid - adress
#     for block in content_blocks:
#         browser.execute_script("arguments[0].scrollIntoView();"  , block )
  
raw_hrefs = []
clean_hrefs = []
 
for block in content_blocks:
    href = block.get_attribute('href')
    raw_hrefs.append(href)

for href in raw_hrefs:
    if 'firm' in str(href):
        clean_hrefs.append(href)

# for c in clean_hrefs:
#     print(c)

# Prints quantity of links found to compare with 2GIS info of branches quantity.


wb_urls = Workbook()
ws_urls = wb_urls.active

row = 1
def write_url(data): 
    ws_urls.cell(row = row, column = 1, value=f'{data}')  
        
for i in clean_hrefs:
    write_url(i)
    row += 1

# SAVE FILE
cwd = os.getcwd()
splitedcwd = os.path.split(cwd)
start = splitedcwd[0]
end = splitedcwd[1]
date = datetime.today().strftime('%Y-%m-%d %H-%M-%S')
filename =  f'{city_name[3]} 2gis_urls'
filetype = '.xlsx'
filepath = os.path.join(start + '/' + end + '/' + filename + ' ' + date + filetype)
wb_urls.save(filepath)
print('\nНайдено',len(clean_hrefs), 'ссылок из' , branch_count)
print(f'\nДанные сохранены в папку "{start}\\{end}\\", файл "{filename} {date}{filetype}"', end='\n\n')
input('Нажмите Enter для выхода.')
wb_urls.close()
browser.quit()  
sys.exit()