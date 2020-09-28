from selenium import webdriver
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import cell
from openpyxl.styles import Alignment #, named_styles
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import psutil
import sys 


print('''
Программа предназначена для получения информации по режиму работы филиалов организации с 2ГИС.
Сначала необходимо выбрать файл ".xlsx", в первом столбце которого 
должны быть прямые ссылки на страницу филиала (точки продаж) начиная с ячейки "А1".
Пример ссылки: "https://2gis.kz/almaty/firm/70000001024622687"

На компьютере должен быть устанолен браузер Firefox, 
потому что программа взаимодействует с ним для поиска дынныъ. 
Чтобы управлять Firefox с помощью программного кода 
используется 'geckodriver' https://github.com/mozilla/geckodriver/releases.
Если браузер не работает, возможно папку с geckodriver.exe необходимо добавить в PATH.
Версия geckodriver.exe должна соответсвовать версии Firefox на компьютере.

Итоговый файл будет сохранен в ту же папку, где была запущена программа.

Во время тестирования обнаружен значительный расход оперативной памяти со стороны браузера Firefox.
При обработке 36-й ссылки из 58 примерный расход памяти браузером составил 2,5 гигабайт.
При обработке 54-й ссылки из 58 примерный расход памяти браузером составил 4 гигабайта.
Если оперативной памяти недостаточно, следует раздробить список ссылок на несколько частей
и отдельно запускать программу для каждой из частей.
''' , end='') 

ram_total = ''
ram_available = ''
ram_percent_available = ''
def ram_info(memory='Текущий объем памяти вашего компьютера:'):
    global ram_total
    global ram_available
    global ram_percent_available
    ram_total = psutil.virtual_memory().total / 1024**3
    ram_available = psutil.virtual_memory().available / 1024**3
    ram_percent_available = psutil.virtual_memory().available * 100 / psutil.virtual_memory().total
    print('-' * 80)
    print(memory)
    print('  Всего памяти (в гигабайтах): ' +  "{:.1f}".format(ram_total )  )
    print('  Свободно памяти (в гигабайтах): ' + "{:.1f}".format(ram_available) )
    print('  Свободно памяти (в процентах): ' + "{:.1f}".format(ram_percent_available))
    print('=' * 80, '\n')

ram_info()
# PROVIDE ".xlsx" FILE WITH 2GIS LINKS IN FIRST COLUMN
# EXAMPLE OF LINK: 'https://2gis.kz/nur_sultan/geo/9570784863367204'
Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
gis_links_file = askopenfilename(filetypes=[("Excel files", ".xlsx")]) # show an "Open" dialog box and return the path to the selected file
gis_links_file = load_workbook(filename = gis_links_file)
gis_links_sheet = gis_links_file.active
row_count = gis_links_sheet.max_row
wbTarget = Workbook()
wsTarget = wbTarget.active

gis_links_to_search = []
# COLLECTING LINKS FROM SOURCE FILE ".xlsx"
print('Собираю ссылки 2ГИС с первого столбца файла ".xlsx"')
for i in range(1, row_count + 1): # '+1' to write the last row
    cell_value = gis_links_sheet.cell(row = i, column = 1).value
    gis_links_to_search.append(cell_value)

print('Начал работу по сбору данных\n\n')
browser = webdriver.Firefox()
street = ''
region = ''
phone = ''
timetable_data = []
headers = ["Регион",'Улица', 'Пон', "Вт", "Ср", "Чт", "Пт", "Сб", "Вс", "Все дни", "Тел", 'Ссылка']
 

# SAVE FILE
def save_exit():
    cwd = os.getcwd()
    splitedcwd = os.path.split(cwd)
    start = splitedcwd[0]
    end = splitedcwd[1]
    date = datetime.today().strftime('%Y-%m-%d %H-%M-%S')
     
    city_name = gis_links_to_search[0].split('/')
    filename =  f'{city_name[3]} 2gis_timetable'
    filetype = '.xlsx'
    filepath = os.path.join(start + '/' + end + '/' + filename + ' ' + date + filetype)
    ram_info('Состояние памяти на момент сохранения:')
    wbTarget.save(filepath)
    wbTarget.close()
    gis_links_file.close()
    browser.quit()

    print(f'\nЗакончил работу.\nДанные сохранены в папку "{start}\\{end}\\", файл "{filename} {date}{filetype}"', end='\n\n')
    input('Нажмите Enter для выхода.')
    sys.exit() # comment this line if you want program to go on


def get_region_street_phone():
    global street
    global region
    global phone
    street = browser.find_element_by_class_name('_er2xx9') 
    street = street.text
    region = browser.find_element_by_class_name('_1p8iqzw') 
    region = region.text
    phone = browser.find_element_by_css_selector("div._b0ke8 > a").get_attribute('href')


def make_clean_list(data_list):
    
    del data_list[:2]
    if 'время работы' in data_list:
        data_list.remove('время работы')  
        if 'обед' in data_list:      
            data_list.remove('обед')    

    return data_list

def find_elements_by_class_name_and_make_list():
    global timetable_data
    timetable_data = browser.find_elements_by_class_name('_18zamfw')[0] 
    timetable_data = timetable_data.text
    timetable_data = timetable_data.split('\n')
    # print(street , '\n', timetable_data,  end='\n' )
    return timetable_data


def prepare_data_for_row_with_lunch(data_list, mon=1, tue=3, wed=5, thu=7, fri=9, sat=-4, sun=-2, all_time=-1):
    global region
    global street
    global phone

    lunch =  data_list[1][11:]

    if mon == False:
        mon = ''
    else:    
        mon = data_list[mon][:11] 
    
    if tue == False:
        tue = ''
    else:
        tue = data_list[tue][:11] 
    
    if wed == False:
        wed = ''
    else:       
        wed = data_list[wed][:11]
    
    if thu == False:
        thu = ''
    else:     
        thu = data_list[thu][:11]
    
    if fri == False:
        fri = ''
    else:    
        fri = data_list[fri][:11]   

    if sat == False:
        sat = ''
    else:      
        sat = data_list[sat]    

    if sun == False:
        sun = ''
    else:       
        sun = data_list[sun]

    if all_time == False:
        all_time = ''
    else:    
        all_time = data_list[all_time]

    data_list = [region, street, mon, tue, wed, thu, fri, sat, sun, all_time + '. Обед будни:' + lunch, phone]
    return data_list


def prepare_data_for_row(data_list, mon=1, tue=3, wed=5, thu=7, fri=9, sat=-4, sun=-2, all_time=-1):
    global region
    global street
    global phone

    if mon == False:
        mon = ''
    else:    
        mon = data_list[mon] 

    if tue == False:
        tue = ''
    else:
        tue = data_list[tue] 

    if wed == False:
        wed = ''
    else:     
        wed = data_list[wed] 

    if thu == False:
        thu = ''
    else:
        thu = data_list[thu] 

    if fri == False:
        fri = ''
    else:
        fri = data_list[fri]

    if sat == False:
        sat = ''
    else:
        sat = data_list[sat]  
    
    if sun == False:
        sun = ''
    else:
        sun = data_list[sun]
   
    if all_time == False:
        all_time = ''
    else:    
        all_time = data_list[all_time]

    data_list = [region, street, mon, tue, wed, thu, fri, sat, sun, all_time, phone] 
    return data_list


def prepare_data_for_excel(data_list):
    '''
    Assigns company working hours to variables. Prepares data for writing to excel spreadsheet.
    
    After [.find_element_by_class_name('_18zamfw')] call not all data is catched. Some working days may be missing.
    That created different cases to work with. 
     
    Function may cause [TypeError: 'NoneType' object is not iterable] in [for loop] if none of conditions apply.

    Order of [if] statements matter. More specific [if] statements must come before less specific.
    '''

    global street
    global region
    global phone

# 15 ['Пн', '07:00–16:00', 'Вт', '07:00–16:00', 'Ср', '07:00–16:00',
#     'Чт', '07:00–16:00', 'Пт', '07:00–16:00', 'Сб', '07:00–13:00',
#     'Вс', '07:30–13:00', 'прием анализов: пн-пт 7:00-15:30; сб 7:00-12:00, вс 7:30-12:00; выдача анализов: пн-пт 7:00-16:00, сб 7:00-13:00, вс 7:30-13:00']
    if (len(data_list) == 15 
    and ('Пн'==  data_list[0] and 'Вт' ==  data_list[2] and 'Ср' ==  data_list[4] 
    and 'Чт' ==  data_list[6] and 'Пт' ==  data_list[8] and 'Сб' ==  data_list[10] 
    and 'Вс' ==  data_list[12])):
        data_list = prepare_data_for_row(data_list, mon=1, tue=3, wed=5, thu=7, fri=9, sat=11, sun=-2, all_time=-1)
        return data_list

# 14 ['Пн', '07:00–19:00', 'Вт', '07:00–19:00', 'Ср', '07:00–19:00', 
#     'Чт', '07:00–19:00', 'Пт', '07:00–19:00', 'Сб', '08:00–12:00', 
#     'Вс', '08:00–12:00']
    if (len(data_list) == 14 
    and ('Пн'==  data_list[0] and 'Вт' ==  data_list[2] and 'Ср' ==  data_list[4] 
    and 'Чт' ==  data_list[6] and 'Пт' ==  data_list[8] and 'Сб' ==  data_list[10] 
    and 'Вс' ==  data_list[-2])):
       data_list = prepare_data_for_row(data_list, mon=1, tue=3, wed=5, thu=7, fri=9, sat=11, sun=-1, all_time=False)
       return data_list
 

# 13 ['Пн', '08:00–18:0013:00–14:00', 'Вт', '08:00–18:0013:00–14:00', 
#     'Чт', '08:00–18:0013:00–14:00', 'Пт', '08:00–18:0013:00–14:00', 'Сб', '——', 
#     'Вс', '——', 'прием анализов: пн-пт 8:00-12:00']
    if (len(data_list) == 13 and  (len(data_list[1]) > 12)
    and ('Пн'==  data_list[0] and 'Вт' ==  data_list[2]   
    and 'Чт' ==  data_list[4] and 'Пт' ==  data_list[6] and 'Сб' ==  data_list[8] 
    and 'Вс' ==  data_list[10])):
        data_list = prepare_data_for_row_with_lunch(data_list, mon=1, tue=3, wed=False, thu=5, fri=7, sat=9, sun=-2, all_time=-1)
        return data_list

# 13 ['Пн', '07:00–15:00', 'Вт', '07:00–15:00', 'Ср', '07:00–15:00',
#     'Чт', '07:00–15:00',                      'Сб', '08:00–15:00', 
#     'Вс', '—', 'прием анализов: пн-пт 7:00-13:00; сб 8:00-12:00']
    if (len(data_list) == 13 
    and ('Пн'==  data_list[0] and 'Вт' ==  data_list[2] and 'Ср' ==  data_list[4] 
    and 'Чт' ==  data_list[6] and                           'Сб' ==  data_list[8] 
    and 'Вс' ==  data_list[10])):
       data_list = prepare_data_for_row(data_list, mon=1, tue=3, wed=5, thu=7, fri=False, sat=9, sun=-2, all_time=-1)
       return data_list

# 13 ['Пн', '08:00–18:00', 'Вт', '08:00–18:00', 'Ср', '08:00–18:00', 
#                          'Пт', '08:00–18:00', 'Сб', '08:00–18:00',
#     'Вс', '—', 'прием анализов: пн-пт 8:00-13:00; сб 8:00-12:00']
    if (len(data_list) == 13 
    and ('Пн'==  data_list[0] and 'Вт' ==  data_list[2] and 'Ср' ==  data_list[4] 
                              and 'Пт' ==  data_list[6] and 'Сб' ==  data_list[8] 
    and 'Вс' ==  data_list[10])):
       data_list = prepare_data_for_row(data_list, mon=1, tue=3, wed=5, thu=False, fri=7, sat=9, sun=-2, all_time=-1)
       return data_list

'''
# 15 ['Пн', '08:00–17:0012:00–13:00', 'Вт', '08:00–17:0012:00–13:00', 'Ср', '08:00–17:0012:00–13:00',
#  'Чт', '08:00–17:0012:00–13:00', 'Пт', '08:00–17:0012:00–13:00', 'Сб', '——', 'Вс', '08:00–14:00—', 
# 'прием анализов: пн-пт 8:00-17:00; вс 8:00-14:00']
    if len(data_list) == 15 and  ((len(data_list[1]) > 12)  and  ('Чт'==  data_list[6] and 'Пт' ==  data_list[8] and 'Вс' ==  data_list[-3])):
        lunch =  data_list[1][11:]
        mon = data_list[1][:11] 
        tue = data_list[3][:11] 
        wed = data_list[5][:11] 
        thu = data_list[7][:11]
        fri = data_list[9][:11]         
        sat = data_list[-4]       
        sun = data_list[-2]
        all_time = data_list[-1]
        data_list = [region, street, mon, tue, wed, thu, fri, sat, sun, all_time + '. Обед будни:' + lunch, phone]
        return data_list
     
'''        

def get_gis_data(gis_link_to_search):
    # if gis_link_to_search.startswith('https://go.2gis'):
    #     browser.get(gis_link_to_search)
    #     browser.find_elements(By.CLASS_NAME, "_18zamfw")[1].click()

    # Scroll if there is Advertisement of company official website
    try:
        browser.get(gis_link_to_search)
        browser.find_element_by_class_name('_5kaapu') # Element of 'Перейти на сайт'
        contents = browser.find_elements_by_class_name('_z3fqkm') # _z3fqkm - arrows
        browser.execute_script("arguments[0].scrollIntoView();"  , contents[-1] )
        # browser.find_elements_by_class_name('_z3fqkm')[1].click()
    except NoSuchElementException:    
        pass

    try:

        global street
        global region
        global phone
        global timetable_data

        
        # Checks if there are two arrows or one
        if len(browser.find_elements_by_class_name('_z3fqkm')) == 1:
            browser.find_elements_by_class_name('_z3fqkm')[0].click()
        else:
            browser.find_elements_by_class_name('_z3fqkm')[1].click() # _z3fqkm - arrow; 
        
        get_region_street_phone()

        timetable_data = browser.find_element_by_class_name('_18zamfw') # _18zamfw - timetable part
        timetable_data = timetable_data.text
        timetable_data = timetable_data.split('\n')
        

        if len(timetable_data) != 1:        
            # No word 'фото' in first div._18zamfw
            timetable_data = make_clean_list(timetable_data)
            print(street, '\n', len(timetable_data),timetable_data, end='\n' )             
            timetable_data = prepare_data_for_excel(timetable_data)
        else:
            # If company has photo the array will only contain text 'фото'             
            # Has word 'фото' in first div._18zamfw
            timetable_data = browser.find_elements_by_class_name('_18zamfw')[1]
            timetable_data = timetable_data.text
            timetable_data = timetable_data.split('\n')
            timetable_data = make_clean_list(timetable_data)
            print(street, '\n', len(timetable_data), timetable_data, end='\n' )
            timetable_data = prepare_data_for_excel(timetable_data)       
          
    except:
        try:
            # Branch temporarily doesn't work -  Филиал временно не работает
            browser.find_element_by_class_name('_1xm5wvm').text # to check if there will be an error
            get_region_street_phone()
            timetable_data = [region, street, 'не работает', 'не работает', 'не работает', 'не работает', 'не работает', 'не работает', 'не работает', 'не работает', phone] 
            print(street, '\n', len(timetable_data), timetable_data, end='\n' )

        except:
            if 'фото' not in find_elements_by_class_name_and_make_list():
                get_region_street_phone()
                timetable_data = browser.find_elements_by_class_name('_18zamfw')[0] 
                timetable_data = timetable_data.text
                timetable_data = timetable_data.split('\n')
                print(street , '\n', timetable_data,  end='\n' )

                # Works everyday             
                if any('Ежедневно' in s for s in timetable_data) and (len(timetable_data) == 3):
                    timetable_data = [region, street, 'Ежедневно', 'Ежедневно', 'Ежедневно', 'Ежедневно', 'Ежедневно', 'Ежедневно', 'Ежедневно', timetable_data[-1], phone]
                elif any('Ежедневно' in s for s in timetable_data) and (len(timetable_data) == 2):
                    timetable_data = [region, street, timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], '', phone]
                else:
                # Works monday to friday
                    timetable_data = make_clean_list(timetable_data)             
                    timetable_data = prepare_data_for_excel(timetable_data)
            else:
                # Works everyday. Found data on second element
                get_region_street_phone()
                timetable_data = browser.find_elements_by_class_name('_18zamfw')[1] 
                timetable_data = timetable_data.text
                timetable_data = timetable_data.split('\n')
                print(street , '\n', timetable_data,  end='\n' )
                timetable_data = [region, street, timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], timetable_data[0], '', phone]
            
        
         
def write_headers(headers):
    col = 1  
    for header in headers:
        wsTarget.cell(row=1, column=(col), value=f'{header}')  
        col += 1  
 

row = 2
def write_row(timetable_data_list):
    col = 1  
    for data in timetable_data_list:
        wsTarget.cell(row = row, column = col, value=f'{data}')  
        col += 1  
    
# MAIN LOOP 
write_headers(headers)


for gis_link in gis_links_to_search:
    try:
        get_gis_data(gis_link)
        write_row(timetable_data)
        wsTarget.cell(row = row, column = len(headers)   , value=f'{gis_link}')
        print(f'Обработано ссылок: {row - 1} из {len(gis_links_to_search)}.\n')
    except Exception as e:
        print('Ошибка:', e)
        print(f'Обработано ссылок: {row - 2} из {len(gis_links_to_search)}.')
        save_exit()    

    row += 1
   
save_exit()

